""" report common"""
#!/usr/bin/python3
# -*- coding: utf-8 -*-
import os
import openpyxl
import openpyxl.styles
from openpyxl.utils import get_column_letter

class SysappBspGeneralReport:
    """ Base class for bsp general case report"""

    def __init__(self,report_file_name,column_headder):
        """Report __init__

        Attributes:
                report_file_name : report file (.xlsx)
        """

        if '/' in report_file_name:
            path,parts = report_file_name.rsplit('/',1)
        else:
            parts = report_file_name
            path = '.'

        if '.' in parts:
            name, _ = parts.rsplit('.', 1)
        else:
            name =   parts

        name = f"{name}.xlsx"

        if not name.startswith('report_'):
            name = 'report_' + name

        report_file_name = path + '/' + name

        alignment = openpyxl.styles.Alignment(horizontal="center",vertical="center")

        if not os.path.exists(report_file_name):
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.row_dimensions[1].height = 50
            self.worksheet.append(column_headder)
            col=1
            while col <= self.worksheet.max_column:
                self.worksheet.column_dimensions[get_column_letter(col)].width = 30
                cell = self.worksheet.cell(1, col)
                cell.alignment = alignment
                cell.font = openpyxl.styles.Font(name="微软雅黑",size=24,bold=True)
                col += 1
        else:
            self.workbook = openpyxl.load_workbook(report_file_name)

        #can use
        self.workbook.save(report_file_name)
        self.current_name = report_file_name
        self.column_headder = column_headder
        self.alignment = alignment
        self.module_index = 0

    def save(self,report_file_rename=None):
        """Report save

        Attributes:
                report_file_rename : renamne report file (.xlsx)
        """
        if report_file_rename is not None:
            self.current_name = report_file_rename

        self.workbook.save(self.current_name)

    def add_sheet_layout(self,sheet_name):
        """Add a sheet into a report file.

        """
        self.worksheet = self.workbook.active

        if len(self.workbook.sheetnames) == 1 and self.workbook.sheetnames[0] == 'Sheet':

            self.worksheet.title = sheet_name
        else:

            new_sheet = self.workbook.copy_worksheet(self.worksheet)
            new_sheet.title = sheet_name
            self.workbook.active = new_sheet
            self.worksheet = self.workbook.active

        self.save()

    def file_test_item(self,test_name,merge_num):
        """Add a sheet into a report file.

        """
        start_cell = 'A2'
        stop_cell_num = 2 + merge_num -1
        stop_cell = 'A' + str(stop_cell_num)
        merge_cell = start_cell + ':' + stop_cell
        worksheet = self.worksheet
        worksheet.merge_cells(merge_cell)

        cell = self.worksheet.cell(2,1)
        cell.alignment = self.alignment
        cell.value = test_name
        self.save()

    def fill_all_result(self,module,test_result:str,detil_info):
        """Fill report file.

        """
        detil_str = ""
        pass_str = "PASS"
        mid_index = module.find('Mid')
        quote_index = module.find("'", mid_index)
        extracted_string = module[(mid_index + 3):quote_index]

        file_cell = self.worksheet.cell(self.module_index + 2, 2)
        file_cell.alignment = self.alignment
        file_cell.value = extracted_string

        file_cell = self.worksheet.cell(self.module_index +2, 3)
        file_cell.alignment = self.alignment
        if test_result == pass_str:
            file_cell.font = openpyxl.styles.Font(color="00FF00")
        else :
            file_cell.font = openpyxl.styles.Font(color="FF0000")

        file_cell.value = test_result

        file_cell = self.worksheet.cell(self.module_index +2, 4)
        file_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        for info in detil_info:
            detil_str += info + "\n"

        file_cell.value = detil_str

        self.module_index += 1
        self.save()
